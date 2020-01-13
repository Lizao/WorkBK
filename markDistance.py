#coding=utf-8
import cv2
import numpy
from ctypes import * 
import struct
import openpyxl

IMGWIDTH = 1280
IMGLENGTH = 800

class RangeInfo(Structure):
    _fields_ = [
        ("fworldx", c_float),
        ("fworldy", c_float),
        ("verDis", c_float),
        ("horDis", c_float)
    ]


AllXmin = []
AllYmin = []
AllXmax = []
AllYmax = []
objNum = []
imgOrder = []

class markDisrtance():
    def __init__(self):
        self.path = '/home/toby/Projects/SVM/SVMData/袁瑞第一批数据/Img_output/'
        self.resultPath = '/home/toby/Projects/SVM/SVMData/袁瑞第一批数据/result/'
        self.imgNum = 51
    
    def readExcel(self):
        file = '/home/toby/Desktop/data.xlsx'
        wb = openpyxl.load_workbook(file)
        sheet = wb['test']
        picture_num = -1
        for row in (range(2, sheet.max_row)): 
            col1 = sheet.cell(row, 1).value
            if col1 != None:
                picture_num = col1
            ymin = IMGLENGTH * sheet.cell(row, 2).value
            xmin = IMGWIDTH * sheet.cell(row, 3).value
            ymax = IMGLENGTH * sheet.cell(row, 4).value
            xmax = IMGWIDTH * sheet.cell(row, 5).value
            AllYmin.append(ymin)
            AllXmin.append(xmin)
            AllYmax.append(ymax)
            AllXmax.append(xmax)
    
    def readTxt(self):
        with open('/home/toby/Projects/SVM/SVMData/袁瑞第一批数据/Img_box.txt') as f1:
            f11 = f1.readlines()

        for i in range(len(f11)):
            objectNum = 0
            content = f11[i][1:-2]
            # print content
            if content.find(".jpg") != -1:
                imgOrder.append(i)
                # print content
            if content.find(".jpg") == -1:
                AllYmin.append(IMGLENGTH * float(content.split(",", 3)[0]))
                AllXmin.append(IMGWIDTH *float(content.split(",", 3)[1]))
                AllYmax.append(IMGLENGTH * float(content.split(",", 3)[2]))
                AllXmax.append(IMGWIDTH *float(content.split(",", 3)[3]))
        for nu in range(1,len(imgOrder)):
            num = imgOrder[nu] - imgOrder[nu-1] - 1
            objNum.append(num)
        objNum.append(len(f11) - imgOrder[-1] -1)
       
    def readVideo(self):
        videoName = "od_test2.mp4"
        cap = cv2.VideoCapture(videoName)
        i = 0
        index = 0
        while i < self.imgNum:
            ret, frame = cap.read()
            frame_name = str(i) + ".jpg"
            # print i
            cv2.imwrite(frame_name, frame)
            cur = cdll.LoadLibrary('./libmonoRanging.so')
            cur.getRange.restype = POINTER(RangeInfo)
            for obj in range (objNum[i]):
                info = cur.getRange(0, c_float(AllXmin[index]), c_float(AllYmin[index]), c_float(AllXmax[index]), c_float(AllYmax[index]))
                text = str(round(info.contents.verDis, 2)) + "m"
                # if info.contents.verDis < 0 :
                #     cv2.circle(frame, (int(AllXmax[index]), int(AllYmax[index])), 10, (0, 0, 255), 5)
                #     print i+1, text, index, AllYmin[index], AllXmin[index], AllYmax[index], AllXmax[index]
                ret = cv2.putText(frame, text, (int(AllXmax[index]) - 15, int(AllYmax[index]) + 7), cv2.FONT_HERSHEY_COMPLEX, 0.3, (0, 255, 0), 1)
                index = index + 1
            frame_name = frame_name[:-4] + ".jpg"
            cv2.imwrite(self.resultPath + frame_name, frame)
            # print index
            i = i + 1
            # print i
    
    # index not from 1
    # def readImg(self):
    #     path = "/home/toby/Projects/SVM/svm/calibrate/monoRanging/img/"
        
    #     i = 1347
    #     index = 0
    #     while (i-1347) < self.imgNum:
    #         name =  str(i) + ".jpg"
    #         img = cv2.imread(path + name)
    #         cur = cdll.LoadLibrary('./libmonoRanging.so')
    #         cur.getRange.restype = POINTER(RangeInfo)
    #         for obj in range (objNum[i-1347]):
    #             info = cur.getRange(0, c_float(AllXmin[index]), c_float(AllYmin[index]), c_float(AllXmax[index]), c_float(AllYmax[index]))
    #             text = str(round(info.contents.verDis, 2)) + "m"
    #             ret = cv2.putText(img, text, (int(AllXmax[index]) - 15, int(AllYmax[index]) + 7), cv2.FONT_HERSHEY_COMPLEX, 0.3, (0, 255, 0), 1)
    #             index = index + 1

    #         # frame_name = frame_name[:-4] + ".jpg"
    #         cv2.imwrite(self.resultPath + name, img)
    #         print index
    #         i = i + 1
    #         # print i
    # index from 1
    def readImg(self):
        path = "/home/toby/Projects/SVM/SVMData/袁瑞第一批数据/Img_output/"
        
        i = 0
        index = 0
        while i < self.imgNum:
            name =  str(i+1) + ".jpg"
            img = cv2.imread(path + name)
            cur = cdll.LoadLibrary('./libmonoRanging.so')
            cur.getRange.restype = POINTER(RangeInfo)
            for obj in range (objNum[i]):
                if i == 49:
                    print AllXmin[index], AllYmin[index], AllXmax[index], AllYmax[index]
                info = cur.getRange(0, c_float(AllXmin[index]), c_float(AllYmin[index]), c_float(AllXmax[index]), c_float(AllYmax[index]))
                text = str(round(info.contents.verDis, 2)) + "m"
                pointUp = '(' + str(round(AllXmin[index], 2)) + "," + str(round(AllYmin[index], 2)) + ')'
                pointDown = '(' + str(round(AllXmax[index], 2)) + "," + str(round(AllYmax[index], 2)) + ')'
                ret0 = cv2.putText(img, text, (int(AllXmax[index]) - 5, int(AllYmax[index]) - 7), cv2.FONT_HERSHEY_COMPLEX, 0.4, (255, 0, 0), 1)
                ret1 = cv2.putText(img, pointUp, (int(AllXmin[index]) - 20, int(AllYmin[index]) + 7), cv2.FONT_HERSHEY_COMPLEX, 0.4, (0, 255, 0), 1)
                ret2 = cv2.putText(img, pointDown, (int(AllXmax[index]) - 30, int(AllYmax[index]) + 17), cv2.FONT_HERSHEY_COMPLEX, 0.4, (0, 0, 255), 1)
                index = index + 1
            # frame_name = frame_name[:-4] + ".jpg"
            cv2.imwrite(self.resultPath + name, img)
            i = i + 1
            # print i

    def img2video(self):
        print "hello"
        fps = 8
        size = (1280, 800)
        video = cv2.VideoWriter("car.avi", cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'), fps, size)
        for i in range(self.imgNum):
        # for i in range(130, 290):
            name = str(i) + ".jpg"
            print name
            img = cv2.imread(self.resultPath + name)
            # cv2.imshow("hi", img)
            # cv2.waitKey()
            video.write(img)
            print i


if __name__ == '__main__':
 hi = markDisrtance()
 # # hi.readExcel()
 hi.readTxt()
 # print "read Txt end"
 hi.readImg()
 # print "read video end"
 # hi.img2video()
 print "img2video end"


 
